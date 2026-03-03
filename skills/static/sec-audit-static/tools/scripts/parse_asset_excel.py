#!/usr/bin/env python3
"""
자산정보 Excel 파싱 스크립트
고객 제공 자산정보 Excel 파일을 JSON으로 변환합니다.

사용법:
    python parse_asset_excel.py <excel_file> [--output <output_file>] [--sheet <sheet_name>]
    python parse_asset_excel.py assets.xlsx
    python parse_asset_excel.py assets.xlsx --output state/task_11_result.json
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl이 설치되어 있지 않습니다.")
    print("  pip install openpyxl")
    sys.exit(1)


# 한/영 헤더 매핑 (한글 헤더 → 영문 키)
HEADER_MAP = {
    # 자산명 / 서비스명
    "자산명": "asset_name",
    "자산 명": "asset_name",
    "시스템명": "asset_name",
    "시스템 명": "asset_name",
    "서비스명": "asset_name",
    "서비스명칭": "asset_name",
    "서비스명칭 [Lv2]": "asset_name",
    "asset_name": "asset_name",
    "asset name": "asset_name",
    "system name": "asset_name",
    "service name": "asset_name",
    # 서비스 상세명
    "서비스명칭 [Lv3]": "service_detail",
    # 서비스 부문/그룹
    "서비스부문": "service_group",
    "서비스군명칭 [Lv1]": "service_group",
    # 자산 유형 (용도 기반: WEB/API/Batch)
    "자산유형": "asset_type",
    "자산 유형": "asset_type",
    "유형": "asset_type",
    "asset_type": "asset_type",
    "asset type": "asset_type",
    "type": "asset_type",
    # 용도 (API, WEB, Batch 등)
    "용도": "purpose",
    "서비스 용도": "purpose",
    "purpose": "purpose",
    # 환경 구분 (개발/상용/QA)
    "구분": "environment",
    # 대내/외
    "대내/외": "exposure",
    "대내/대외": "exposure",
    # 도메인 / URL
    "도메인": "domain",
    "URL": "domain",
    "url": "domain",
    "domain": "domain",
    "서비스 URL": "domain",
    "서비스URL": "domain",
    "접속 URL": "domain",
    # Public IP
    "IP": "ip",
    "ip": "ip",
    "IP주소": "ip",
    "IP 주소": "ip",
    "ip address": "ip",
    "ip_address": "ip",
    "서버IP": "ip",
    "서버 IP": "ip",
    "Public IP": "ip",
    # Private IP
    "Private IP": "private_ip",
    # 포트
    "포트": "ports",
    "port": "ports",
    "Port": "ports",
    "ports": "ports",
    "서비스포트": "ports",
    "서비스 포트": "ports",
    # 중요도
    "중요도": "criticality",
    "등급": "criticality",
    "위험도": "criticality",
    "criticality": "criticality",
    "severity": "criticality",
    "importance": "criticality",
    # 기술 스택
    "기술스택": "tech_stack",
    "기술 스택": "tech_stack",
    "사용기술": "tech_stack",
    "개발언어": "tech_stack",
    "개발 언어": "tech_stack",
    "프레임워크": "tech_stack",
    "tech_stack": "tech_stack",
    "tech stack": "tech_stack",
    "technology": "tech_stack",
    "framework": "tech_stack",
    "language": "tech_stack",
    # 언어/WAS 정보
    "언어종류/버전": "language_version",
    "WAS 종류/버전": "was_version",
    "로깅툴/버전": "logging_tool",
    # 담당자
    "담당자": "owner",
    "관리자": "owner",
    "owner": "owner",
    "manager": "owner",
    "개발담당자1 이름": "dev_owner",
    "사업담당자1 이름": "biz_owner",
    # 상태
    "상태": "status",
    # 인증/결제
    "인증기능": "has_auth",
    "결제기능": "has_payment",
    # 번호
    "No": "no",
    # 소스 저장소
    "Repository 주소": "repository_url",
    "Branch 명": "branch",
    # 비고
    "비고": "notes",
    "설명": "notes",
    "notes": "notes",
    "description": "notes",
    "remarks": "notes",
    # OS
    "OS": "os",
    "os": "os",
    "운영체제": "os",
    "운영 체제": "os",
}


def normalize_header(raw: str) -> str | None:
    """헤더 문자열을 정규화하여 영문 키로 변환합니다."""
    if not raw:
        return None
    # 줄바꿈/탭을 공백으로 치환 후 연속 공백 정리
    import re
    cleaned = re.sub(r"[\n\r\t]+", " ", str(raw)).strip()
    result = HEADER_MAP.get(cleaned)
    if result:
        return result
    # 줄바꿈 제거 버전으로도 시도
    collapsed = re.sub(r"\s+", " ", cleaned)
    return HEADER_MAP.get(collapsed)


def parse_tech_stack(value) -> list[str]:
    """기술 스택 문자열을 리스트로 변환합니다."""
    if not value:
        return []
    text = str(value).strip()
    for sep in [",", "/", ";", "\n"]:
        if sep in text:
            return [item.strip() for item in text.split(sep) if item.strip()]
    return [text]


def parse_ports(value) -> list[int]:
    """포트 문자열을 정수 리스트로 변환합니다."""
    if not value:
        return []
    text = str(value).strip()
    ports = []
    for part in text.replace(";", ",").replace("/", ",").split(","):
        part = part.strip()
        try:
            ports.append(int(part))
        except ValueError:
            continue
    return ports


def parse_excel(file_path: Path, sheet_name: str | None = None) -> list[dict]:
    """Excel 파일을 파싱하여 자산 목록을 반환합니다."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: 시트 '{sheet_name}'를 찾을 수 없습니다. 사용 가능: {wb.sheetnames}")
            ws = wb.active
        else:
            ws = wb[sheet_name]
    else:
        ws = wb.active

    print(f"시트: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 헤더 행 탐색 (최대 10행까지 확인)
    header_row_idx = None
    header_map = {}
    for idx, row in enumerate(rows[:10]):
        mapped = {}
        for col_idx, cell in enumerate(row):
            key = normalize_header(cell)
            if key:
                mapped[col_idx] = key
        # 최소 2개 이상의 인식된 헤더가 있으면 헤더 행으로 판단
        if len(mapped) >= 2:
            header_row_idx = idx
            header_map = mapped
            break

    if header_row_idx is None:
        print("Warning: 인식 가능한 헤더를 찾을 수 없습니다.")
        print(f"  첫 번째 행: {rows[0] if rows else '(비어있음)'}")
        return []

    recognized = {v for v in header_map.values()}
    print(f"인식된 헤더: {recognized}")

    # 데이터 행 파싱
    assets = []
    for row in rows[header_row_idx + 1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        asset = {}
        for col_idx, key in header_map.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    if key == "tech_stack":
                        asset[key] = parse_tech_stack(value)
                    elif key == "ports":
                        asset[key] = parse_ports(value)
                    else:
                        asset[key] = str(value).strip()

        # 예제/헤더 행 건너뛰기
        no_val = str(asset.get("no", "")).strip().lower()
        if no_val.startswith("ex") or no_val == "no":
            continue
        # 값이 다른 헤더명과 동일한 행 건너뛰기 (2차 헤더 행)
        purpose_val = str(asset.get("purpose", "")).strip()
        if purpose_val in ("용도", "purpose"):
            continue

        # asset_name 또는 service_detail이 있는 행만 포함
        name = asset.get("asset_name") or asset.get("service_detail")
        if name:
            # asset_name이 없으면 service_detail로 대체
            if not asset.get("asset_name") and asset.get("service_detail"):
                asset["asset_name"] = asset["service_detail"]
            # 내부 필터용 필드 제거
            asset.pop("no", None)
            assets.append(asset)

    wb.close()
    return assets


def build_task_output(
    assets: list[dict],
    source_file: str,
    source_repo_url: str,
    source_repo_path: str,
    source_modules: list[str],
) -> dict:
    """파싱 결과를 task_output_schema.json 형식으로 구성합니다."""
    return {
        "task_id": "1-1",
        "status": "completed",
        "findings": assets,
        "metadata": {
            "source_repo_url": source_repo_url,
            "source_repo_path": source_repo_path,
            "source_modules": source_modules,
            "source_file": source_file,
            "total_assets": len(assets),
            "parse_method": "openpyxl",
        },
        "executed_at": datetime.now(timezone.utc).isoformat(),
        "claude_session": "",
    }


def main():
    parser = argparse.ArgumentParser(description="자산정보 Excel → JSON 변환 도구")
    parser.add_argument("excel_file", help="입력 Excel 파일 경로")
    parser.add_argument(
        "--source-repo-url",
        required=True,
        help="진단 대상 레포 URL (예: http://code.example.com/projects/PROJ/repos/repo/)",
    )
    parser.add_argument(
        "--source-repo-path",
        required=True,
        help="로컬 레포 경로 (예: /Users/.../Downloads/repo)",
    )
    parser.add_argument(
        "--source-modules",
        required=True,
        help="진단 대상 모듈/서브프로젝트 (comma-separated)",
    )
    parser.add_argument(
        "--output", "-o",
        help="출력 JSON 파일 경로 (기본: state/task_11_result.json)",
        default="state/task_11_result.json",
    )
    parser.add_argument(
        "--sheet", "-s",
        help="파싱할 시트 이름 (기본: 첫 번째 시트)",
        default=None,
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: 파일을 찾을 수 없습니다: {args.excel_file}")
        sys.exit(1)

    print(f"파싱 대상: {excel_path}")

    assets = parse_excel(excel_path, args.sheet)
    if not assets:
        print("Warning: 파싱된 자산이 없습니다.")
        sys.exit(1)

    modules = [m.strip() for m in args.source_modules.split(",") if m.strip()]
    if not modules:
        print("Error: --source-modules 값이 비어 있습니다.")
        sys.exit(1)

    result = build_task_output(
        assets,
        str(excel_path.name),
        args.source_repo_url,
        args.source_repo_path,
        modules,
    )

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n변환 완료: {output_path}")
    print(f"  - 총 자산 수: {len(assets)}")
    for i, asset in enumerate(assets[:5]):
        print(f"  - [{i+1}] {asset.get('asset_name', 'N/A')}")
    if len(assets) > 5:
        print(f"  - ... 외 {len(assets) - 5}건")


if __name__ == "__main__":
    main()
